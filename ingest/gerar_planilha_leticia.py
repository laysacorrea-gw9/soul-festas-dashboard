"""Gera o Excel de apropriacoes a corrigir pra Leticia, ja com as decisoes da
reuniao 16/06/2026 (Diesel->Evento, Comissao->Casa, Estorno rescisao->Evento) e
EXCLUINDO Decoracao (decoracao de cartao sem evento fica como Casa, por decisao da Laysa).

Sinaliza quando o Centro de Custo do lancamento contradiz a natureza do servico.
"""
from pathlib import Path
import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.stdout.reconfigure(encoding="utf-8")
ROOT = Path(__file__).parent

CASA_CC = ["Administrativo", "DP", "Comercial"]
EV_CC = ["Eventos", "DEGUSTAÇÃO"]
# Decoracao fica de fora: card sem evento identificado -> Casa (decisao Laysa 16/06)
EXCLUIR = {"DECORAÇÃO", "EXTRAS DE DECORAÇÃO"}

ANO = 2026
DEST = r"C:\Users\Laysa\OneDrive\APAGAR\SOUL\Apropriacoes_Soul_2026_para_Leticia.xlsx"


def main():
    pg = pd.read_csv(ROOT / "data_out" / "contas_pagar_final.csv", low_memory=False)
    nat = pd.read_csv(ROOT / "de_para_natureza_servico.csv")
    nat_map = dict(zip(nat["servico"].str.strip().str.upper(), nat["natureza"]))

    pg["data_ref"] = pd.to_datetime(pg["data_ref"], errors="coerce")
    d = pg[pg["data_ref"].dt.year == ANO].copy()
    d["natureza"] = d["Serviço"].astype(str).str.strip().str.upper().map(nat_map).fillna("CASA")
    d["bloco"] = d["C. Custo"].apply(lambda x: "CASA" if x in CASA_CC else ("EVENTO" if x in EV_CC else "OUTRO"))

    # Tipo A: natureza EVENTO mas C.Custo Casa
    tA = d[(d["natureza"] == "EVENTO") & (d["bloco"] == "CASA") & (~d["Serviço"].astype(str).str.upper().isin(EXCLUIR))].copy()
    tA["Tipo"] = "1 - Evento lancado na Casa"
    tA["Correcao"] = "Mudar Centro de Custo p/ EVENTOS + vincular Projeto"
    # Tipo B: natureza CASA mas C.Custo Evento
    tB = d[(d["natureza"] == "CASA") & (d["bloco"] == "EVENTO")].copy()
    tB["Tipo"] = "2 - Casa lancado no Evento"
    tB["Correcao"] = "Mudar Centro de Custo p/ ADMINISTRATIVO (Casa)"
    sus = pd.concat([tA, tB])
    sus["Descrição"] = sus["Descrição"].fillna("")
    sus["Projeto"] = sus["Projeto"].fillna("(vazio)").replace("nan", "(vazio)")
    sus = sus.sort_values(["Tipo", "valor_ref"], ascending=[True, False])

    GOLD, DARK, WHITE, L1, L2 = "C9A227", "1F2A44", "FFFFFF", "FFF6E0", "EAF2FB"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    wb = Workbook()

    # ---- Resumo ----
    ws = wb.active
    ws.title = "Resumo"
    ws.sheet_view.showGridLines = False
    ws["B2"] = "Soul Eventos — Lançamentos para corrigir no SGE (2026)"
    ws["B2"].font = Font(bold=True, size=16, color=DARK)
    ws["B3"] = "Decisões da reunião de 16/06/2026 (Laysa + Letícia)"
    ws["B3"].font = Font(size=11, italic=True, color="666666")
    tot = sus["valor_ref"].sum()
    ws["B5"] = f"Total: {len(sus)} lançamentos  ·  R$ {tot:,.0f}".replace(",", ".")
    ws["B5"].font = Font(bold=True, size=12, color=DARK)

    res = sus.groupby(["Tipo", "Serviço"])["valor_ref"].agg(["count", "sum"]).reset_index().sort_values(["Tipo", "sum"], ascending=[True, False])
    r0 = 7
    for j, h in enumerate(["Tipo", "Serviço", "Qtd", "Valor (R$)"]):
        c = ws.cell(row=r0, column=2 + j, value=h)
        c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=DARK)
        c.alignment = Alignment(horizontal="center"); c.border = border
    for i, (_, row) in enumerate(res.iterrows()):
        rr = r0 + 1 + i
        fill = L1 if row["Tipo"].startswith("1") else L2
        for j, v in enumerate([row["Tipo"], row["Serviço"], int(row["count"]), float(row["sum"])]):
            c = ws.cell(row=rr, column=2 + j, value=v); c.fill = PatternFill("solid", fgColor=fill); c.border = border
            if j == 3:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right")
            if j == 2:
                c.alignment = Alignment(horizontal="center")
    for col, w in zip("BCDE", [26, 30, 10, 16]):
        ws.column_dimensions[col].width = w
    ws.column_dimensions["A"].width = 3

    # ---- Lista ----
    ws2 = wb.create_sheet("Lista para corrigir")
    ws2.sheet_view.showGridLines = False
    cols = [("Data", 12), ("Fornecedor", 32), ("Descrição", 32), ("Serviço", 22),
            ("Valor (R$)", 13), ("C.Custo ATUAL", 16), ("Projeto ATUAL", 14),
            ("O que corrigir", 42), ("Feito?", 9)]
    for j, (h, w) in enumerate(cols):
        c = ws2.cell(row=1, column=1 + j, value=h)
        c.font = Font(bold=True, color=WHITE); c.fill = PatternFill("solid", fgColor=GOLD)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); c.border = border
        ws2.column_dimensions[get_column_letter(1 + j)].width = w
    for i, (_, r) in enumerate(sus.iterrows()):
        rr = i + 2
        fill = L1 if r["Tipo"].startswith("1") else L2
        vals = [r["data_ref"].strftime("%d/%m/%Y") if pd.notna(r["data_ref"]) else "",
                str(r["Fornecedor"])[:60], str(r["Descrição"])[:80], r["Serviço"],
                float(r["valor_ref"]), r["C. Custo"], str(r["Projeto"]), r["Correcao"], ""]
        for j, v in enumerate(vals):
            c = ws2.cell(row=rr, column=1 + j, value=v); c.fill = PatternFill("solid", fgColor=fill); c.border = border
            c.alignment = Alignment(vertical="center", wrap_text=(j in (2, 7)))
            if j == 4:
                c.number_format = "#,##0.00"; c.alignment = Alignment(horizontal="right", vertical="center")
    ws2.freeze_panes = "A2"; ws2.auto_filter.ref = f"A1:I{len(sus)+1}"; ws2.row_dimensions[1].height = 30

    # ---- Observacoes / procedimentos ----
    ws3 = wb.create_sheet("Observações")
    ws3.sheet_view.showGridLines = False
    ws3.column_dimensions["B"].width = 100
    notas = [
        ("Procedimentos combinados na reunião (16/06/2026)", True),
        ("", False),
        ("1) FATURAS DE CARTÃO — evitar duplicidade (causou caixa negativo em maio):", True),
        ("   • Quando a fatura está lançada CHEIA e também com os itens discriminados, o sistema soma duas vezes.", False),
        ("   • Ao pagar a fatura, lançar o VALOR DE PAGAMENTO = R$ 0,01 para o sistema não contabilizar o valor cheio.", False),
        ("   • Conferir Bradesco (~R$ 14 mil) e demais faturas a partir de março/2026. Só os itens discriminados devem somar.", False),
        ("", False),
        ("2) DIESEL (Velp/Help Transportes): mudar de Despesa Fixa para DESPESA COM EVENTO (tem OS/projeto).", False),
        ("3) COMISSÃO DE VENDAS: mudar Centro de Custo de Evento para ADMINISTRATIVO (Casa).", False),
        ("4) ESTORNO DE RESCISÃO: é cancelamento de evento — manter como DESPESA COM EVENTO (está correto).", False),
        ("5) MANUTENÇÃO E CONSERVAÇÃO: caso a caso — rotineira = Casa; específica de um evento = Evento.", False),
        ("6) BOMBEIRO: taxa anual = Casa.", False),
        ("7) DECORAÇÃO comprada no cartão sem saber o evento: manter como Administrativo (Casa). Quando souber o evento, vincular o projeto.", False),
        ("", False),
        ("Dica: dá pra corrigir vários de uma vez alterando o CADASTRO DO FORNECEDOR (centro de custo/categoria), que atualiza todos os lançamentos dele.", True),
        ("Se algo for muito trabalhoso de corrigir retroativo, sinaliza pra Laysa — o foco é acertar os novos lançamentos daqui pra frente.", False),
    ]
    for i, (txt, bold) in enumerate(notas):
        c = ws3.cell(row=2 + i, column=2, value=txt)
        c.font = Font(bold=bold, size=12 if bold and i == 0 else 11, color=DARK)
        c.alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(DEST)
    print(f"Excel salvo: {DEST}")
    print(f"Tipo 1 (Evento->Casa): {len(tA)} | Tipo 2 (Casa->Evento): {len(tB)} | Total: {len(sus)} | R$ {tot:,.0f}")
    print("\nPor servico:")
    print(res.to_string(index=False))


if __name__ == "__main__":
    main()
